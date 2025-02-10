import openpyxl
import os

# Ruta del CSV de entrada
csv_path = r"C:\Users\gonza\Desktop\Nueva carpeta\Whatsapp-Contacts.csv"

# Ruta del Excel de salida
excel_path = r"C:\Users\gonza\Desktop\Nueva carpeta\numeros_peru_filtrados.xlsx"

def es_numero_peruano_valido(num_str):
    """
    Verifica si el número:
    - Empieza con '51'
    - Tiene exactamente 11 dígitos
    - Son todos dígitos
    """
    return (
        len(num_str) == 11 and
        num_str.startswith("51") and
        num_str.isdigit()
    )

def formatear_numero_peru(num_str):
    """
    Recibe un string con 11 dígitos que comienzan en '51'.
    Retorna el formato +51 999 999 999
    """
    # num_str es algo como '51991735012'
    codigo_pais = num_str[:2]  # '51'
    parte1 = num_str[2:5]      # Ej: '991'
    parte2 = num_str[5:8]      # Ej: '735'
    parte3 = num_str[8:11]     # Ej: '012'
    return f"+{codigo_pais} {parte1} {parte2} {parte3}"

def leer_numeros_manualmente(ruta_csv):
    """
    Lee el archivo línea por línea y extrae el valor
    numérico del primer 'campo', ignorando el resto.
    """
    numeros = []
    with open(ruta_csv, mode="r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            # Quita comillas de inicio/fin si las tiene
            if line.startswith('"') and line.endswith('"'):
                line = line[1:-1]
            
            # Ahora partimos por la primera coma.
            # Queremos el primer "trozo" antes de la primera coma
            # (porque ahí está el número).
            # Por ejemplo, algo como: 51991735012,"""","""" -> '51991735012,"""",""""'
            parts = line.split(",", 1)  # solo dividimos 1 vez
            numero_str = parts[0].strip()
            
            # quita comillas internas
            numero_str = numero_str.replace('"', '').replace("'", "").strip()
            
            # A veces puede haber una coma final
            if numero_str.endswith(","):
                numero_str = numero_str[:-1]
            
            if numero_str:
                numeros.append(numero_str)
    return numeros

def main():
    # 1. Leer números del archivo
    todos_los_numeros = leer_numeros_manualmente(csv_path)
    
    # 2. Filtrar solo los peruanos válidos
    numeros_peru = [num for num in todos_los_numeros if es_numero_peruano_valido(num)]
    
    # 3. Eliminar duplicados
    numeros_peru_unicos = list(set(numeros_peru))
    
    # 4. Formatear
    numeros_formateados = [formatear_numero_peru(num) for num in numeros_peru_unicos]
    
    # 5. Crear Excel y guardar
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Peru (+51)"
    
    ws.cell(row=1, column=1).value = "Número Formateado"
    
    for idx, numero in enumerate(numeros_formateados, start=2):
        ws.cell(row=idx, column=1).value = numero
    
    wb.save(excel_path)
    print("Proceso completado. Archivo creado en:", excel_path)

if __name__ == "__main__":
    main()
