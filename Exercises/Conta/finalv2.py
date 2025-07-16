#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
ANTARES S.A.C. – Proyecciones Feb-Mar 2015
Crea/actualiza el libro Excel añadiendo hojas con formato profesional y
fórmulas dinámicas.

❑ Ventas                 ❑ Compras_COGS          ❑ Gastos_Operativos
❑ Estado_Resultados_Proj ❑ Flujo_Caja_Proj       ❑ Balance_31-03-2015
❑ Ratios_Decisiones
"""

from pathlib import Path
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import (
    Font, Alignment, PatternFill, Border, Side, NamedStyle, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

# ═══════════════════════ 1. PARÁMETROS ─ Premisas editables ════════════════════
PCT_CASH_COLLECT   = 0.68
PCT_CREDIT_COLLECT = 0.32
PCT_COGS           = 0.90
PCT_INV_FINAL      = 0.20
PCT_PURCH_CASH     = 0.75
PCT_PURCH_CREDIT   = 0.25
PCT_COMMISSION     = 0.02
IGV_RATE           = 0.18
IR_ADV_RATE        = 0.015
PARTICIP_RATE      = 0.05
IR_RATE            = 0.30

SALES = {"Feb": 1_350_000, "Mar": 1_689_000}
MONTHS = list(SALES)

INV_FINAL_JAN      = 26_400
CASH_INITIAL_FEB   = 65_347
DEBT_INITIAL_TOTAL = 41_454.93 + 102_497.31
DEBT_PRINCIPAL_PAY = 0.40 * DEBT_INITIAL_TOTAL
DEBT_INTEREST_FEB  = 3_600

SUELDOS               = {"Feb": 19_000, "Mar": 18_500}
ALQUILER              = 2_100
DEPRECIACION_MENSUAL  = 1_666.66
AMORT_SEGUROS_MENSUAL = 2_000

# ═══════════════════════ 2. UTILIDADES DE FORMATO ══════════════════════════════
THIN   = Side(style="thin", color="9C9C9C")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

def col_widths(ws, widths):
    for idx, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = w

def header_style(ws, row, fill="BDD7EE"):
    for c in ws[row]:
        c.font      = Font(bold=True)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.fill      = PatternFill("solid", fgColor=fill)
        c.border    = BORDER
    ws.row_dimensions[row].height = 20

def body_style(ws, start_row):
    for r in ws.iter_rows(min_row=start_row):
        for c in r:
            if isinstance(c.value, (int, float)):
                c.number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1
            c.border    = BORDER
            c.alignment = Alignment(
                horizontal="right" if isinstance(c.value, (int,float)) else "left",
                vertical="center"
            )

def add_title(ws, titulo, subtitulo, moneda="S/.", row=1):
    ws.merge_cells(start_row=row,     start_column=1, end_row=row,     end_column=6)
    ws.merge_cells(start_row=row+1,   start_column=1, end_row=row+1,   end_column=6)
    ws.cell(row,   1, f"ANTARES S.A.C. – {titulo}").font = Font(bold=True, size=14)
    ws.cell(row+1, 1, f"(Expresado en {moneda})   {subtitulo}").font = Font(bold=True, size=12)
    ws.row_dimensions[row].height   = 22
    ws.row_dimensions[row+1].height = 18

def df_to_sheet(wb, name, df, title, subtitle):
    if name in wb.sheetnames:
        del wb[name]
    ws = wb.create_sheet(name)
    add_title(ws, title, subtitle)
    start = 4  # fila donde inicia la tabla
    for r_i, row in enumerate(dataframe_to_rows(df, index=False, header=True), start):
        ws.append(row)
        if r_i == start:           # encabezado
            header_style(ws, r_i)
        elif r_i > start:
            ws.row_dimensions[r_i].height = 18
    body_style(ws, start+1)
    ws.freeze_panes = ws["A5"]
    col_widths(ws, [12, 18, 18, 18, 18, 18])
    return ws

# ═══════════════════════ 3. DATAFRAMES + FÓRMULAS ══════════════════════════════
# 3.1 Ventas y cobranzas
ventas_df = pd.DataFrame({
    "Mes": MONTHS,
    "Ventas": [SALES[m] for m in MONTHS],
    "Cobro Contado": [None]*len(MONTHS),
    "Cobro 30 días": [None]*len(MONTHS)
})

# 3.2 Compras & COGS
compras_df = pd.DataFrame({
    "Mes": MONTHS,
    "Inv. Inicial":  None,
    "COGS":          None,
    "Inv. Final":    None,
    "Compras":       None,
    "Pago Contado":  None,
    "Pago 60 días":  None
})

# 3.3 Gastos operativos
gastos_df = pd.DataFrame({
    "Mes": MONTHS,
    "Sueldos":        [SUELDOS[m] for m in MONTHS],
    "Alquiler":       ALQUILER,
    "Comisiones":     None,
    "Depreciación":   DEPRECIACION_MENSUAL,
    "Amort. Seguros": AMORT_SEGUROS_MENSUAL,
    "Gastos Efectivo": None,
    "Gastos No Efect.": None
})

# 3.4 Estado de resultados
pl_df = pd.DataFrame({
    "Mes": MONTHS,
    "Ventas":             None,
    "Costo Ventas":       None,
    "Utilidad Bruta":     None,
    "Gastos Operativos":  None,
    "Utilidad Operativa": None,
    "Gastos Financieros": None,
    "Resultado Antes Part":None,
    "Partic. Trab (5%)":  None,
    "Util. Antes IR":     None,
    "Impuesto (30%)":     None,
    "Utilidad Neta":      None
})

# 3.5 Flujo de caja
flujo_df = pd.DataFrame({
    "Mes": MONTHS,
    "Efectivo Inicial":  None,
    "Entradas":          None,
    "Salidas":           None,
    "Efectivo Final":    None
})

# ═══════════════════════ 4. ESCRITURA DE ARCHIVO ═══════════════════════════════
book_path = Path(r"D:\FOREX\MT5_1\MQL5\Experts\Projects\Sharing-Projects\Misc\Exercises\Conta\ANTARES_2014_profesional.xlsx")
wb = load_workbook(book_path)

# 4.1 Hoja VENTAS con fórmulas
ws_v = df_to_sheet(wb, "Ventas", ventas_df,
                   "Presupuesto de Ventas y Cobranza", "Feb – Mar 2015")
for i, _ in enumerate(MONTHS, start=5):
    ws_v[f"C{i}"] = f"=ROUND($B{i}*{PCT_CASH_COLLECT},2)"
    ws_v[f"D{i}"] = f"=ROUND($B{i}*{PCT_CREDIT_COLLECT},2)"
tot_row = len(MONTHS) + 6
ws_v[f"A{tot_row}"], ws_v[f"A{tot_row}"].font = "TOTAL", Font(bold=True)
for col in "BCD":
    ws_v[f"{col}{tot_row}"] = f"=SUM({col}5:{col}{tot_row-1})"
    ws_v[f"{col}{tot_row}"].font = Font(bold=True)

# 4.2 Hoja COMPRAS con fórmulas
ws_c = df_to_sheet(wb, "Compras_COGS", compras_df,
                   "Presupuesto de Compras & COGS", "Feb – Mar 2015")
for idx, _ in enumerate(MONTHS, start=5):
    prev_inv = INV_FINAL_JAN if idx == 5 else f"C{idx-1}"
    ws_c[f"B{idx}"] = prev_inv
    ws_c[f"C{idx}"] = f"=ROUND(Ventas!$B{idx}*{PCT_COGS},2)"
    ws_c[f"D{idx}"] = f"=ROUND($C{idx}*{PCT_INV_FINAL},2)"
    ws_c[f"E{idx}"] = f"=ROUND($C{idx}+$D{idx}-$B{idx},2)"
    ws_c[f"F{idx}"] = f"=ROUND($E{idx}*{PCT_PURCH_CASH},2)"
    ws_c[f"G{idx}"] = f"=ROUND($E{idx}*{PCT_PURCH_CREDIT},2)"

# 4.3 Hoja GASTOS OPERATIVOS con fórmulas
ws_g = df_to_sheet(wb, "Gastos_Operativos", gastos_df,
                   "Presupuesto de Gastos Operativos", "Feb – Mar 2015")
for idx, _ in enumerate(MONTHS, start=5):
    ws_g[f"D{idx}"] = f"=ROUND(Ventas!$B{idx}*{PCT_COMMISSION},2)"
    ws_g[f"G{idx}"] = f"=$B{idx}+$C{idx}+$D{idx}"
    ws_g[f"H{idx}"] = f"=$E{idx}+$F{idx}"

# 4.4 Estado de Resultados (100 % fórmulas)
ws_p = df_to_sheet(wb, "Estado_Resultados_Proj", pl_df,
                   "Estado de Resultados Proyectado", "Feb – Mar 2015")
for idx, _ in enumerate(MONTHS, start=5):
    ws_p[f"B{idx}"] = f"=Ventas!$B{idx}"
    ws_p[f"C{idx}"] = f"=-Compras_COGS!$C{idx}"
    ws_p[f"D{idx}"] = f"=$B{idx}+$C{idx}"
    ws_p[f"E{idx}"] = f"=-Gastos_Operativos!$G{idx}-Gastos_Operativos!$H{idx}"
    ws_p[f"F{idx}"] = f"=$D{idx}+$E{idx}"
    intg = -DEBT_INTEREST_FEB if idx == 5 else 0
    ws_p[f"G{idx}"] = intg
    ws_p[f"H{idx}"] = f"=$F{idx}+$G{idx}"
    ws_p[f"I{idx}"] = f"=ROUND($H{idx}*{-PARTICIP_RATE},2)"
    ws_p[f"J{idx}"] = f"=$H{idx}+$I{idx}"
    ws_p[f"K{idx}"] = f"=ROUND($J{idx}*{-IR_RATE},2)"
    ws_p[f"L{idx}"] = f"=$J{idx}+$K{idx}"

# 4.5 Flujo de caja
flujo_sheet = df_to_sheet(wb, "Flujo_Caja_Proj", flujo_df,
                          "Flujo de Caja Proyectado", "Feb – Mar 2015")
for idx, _ in enumerate(MONTHS, start=5):
    if idx == 5:
        flujo_sheet[f"B{idx}"] = CASH_INITIAL_FEB
    else:
        flujo_sheet[f"B{idx}"] = f"=E{idx-1}"
    flujo_sheet[f"C{idx}"] = (
        f"=Ventas!$C{idx}"
        + (f"+Ventas!$D{idx-1}" if idx == 6 else "")
    )
    if idx == 5:
        flujo_sheet[f"D{idx}"] = (
            f"=Compras_COGS!$F{idx}+Gastos_Operativos!$G{idx}"
            f"+Ratios_Decisiones!$B$10"
        )
    else:
        flujo_sheet[f"D{idx}"] = (
            f"=Compras_COGS!$F{idx}+Gastos_Operativos!$G{idx}"
        )
    flujo_sheet[f"E{idx}"] = f"=$B{idx}+$C{idx}-$D{idx}"

# 4.6 Balance 31-03-2015  (números planos y fórmulas)
bal_activo = {
    "Efectivo":          f"=Flujo_Caja_Proj!$E6",
    "Ctas por Cobrar":   f"=Ventas!$D6",
    "Inventario":        f"=Compras_COGS!$D6",
    "Pagos Anticipados": 2_843.95 - 2*AMORT_SEGUROS_MENSUAL,
    "PPE Neta":          978_257.29 - 2*DEPRECIACION_MENSUAL
}
bal_pasivo = {
    "Ctas por Pagar":     f"=Compras_COGS!$G5+Compras_COGS!$G6",
    "Tributos x Pagar":   f"=Ventas!$B6*{IGV_RATE+IR_ADV_RATE}",
    "Oblig. Financieras": DEBT_INITIAL_TOTAL - DEBT_PRINCIPAL_PAY
}
# Definimos la expresión de utilidad retenida sin '=' para luego reutilizarla
util_reten = "Estado_Resultados_Proj!$L5+Estado_Resultados_Proj!$L6"
patrimonio = 870_740.41

balance_df = pd.concat([
    pd.DataFrame(list(bal_activo.items()), columns=["Cuenta","Monto"]),
    pd.DataFrame(list(bal_pasivo.items()), columns=["Cuenta","Monto"]),
    pd.DataFrame([["Patrimonio", None]],        columns=["Cuenta","Monto"])
])

ws_b = df_to_sheet(wb, "Balance_31-03-2015", balance_df,
                   "Estado de Situación Financiera", "Al 31-03-2015")
for r in range(5, 10):
    ws_b[f"B{r}"].number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1
ws_b["B10"] = f"={util_reten}"      # Utilidad retenida
ws_b["B11"] = "=SUM(B5:B9)"         # Patrimonio (sumario)

# 4.7 Ratios + decisiones (hoja con nombre sin espacios, no precisa comillas)
ratios_df = pd.DataFrame({
    "Ratio": [
        "Razón Corriente",
        "Razón Rápida",
        "Endeudamiento",
        "ROE (Feb-Mar)",
        "Margen Neto"
    ],
    "Fórmula": [
        # Se envuelve todo el nombre de la hoja Balance_31-03-2015 entre comillas simples
        "='Balance_31-03-2015'!$B$5/'Balance_31-03-2015'!$B$7",
        "=('Balance_31-03-2015'!$B$5+'Balance_31-03-2015'!$B$6)/'Balance_31-03-2015'!$B$7",
        "='Balance_31-03-2015'!$B$7/('Balance_31-03-2015'!$B$7+'Balance_31-03-2015'!$B$9)",
        f"=({util_reten})/'Balance_31-03-2015'!$B$9",
        f"=({util_reten})/SUM(Ventas!$B$5:Ventas!$B$6)"
    ]
})

ws_r = df_to_sheet(wb, "Ratios_Decisiones", ratios_df,
                   "Ratios Financieros Proyectados", "31-03-2015")
for idx in range(5, 5 + len(ratios_df)):
    ws_r[f"B{idx}"] = ratios_df.at[idx-5, "Fórmula"]
body_style(ws_r, 5)

ws_d = df_to_sheet(wb, "Decisiones_Clave",
                   pd.DataFrame({"Decisiones Clave":[
                       "Extender pago a proveedores a 90 días",
                       "Descuento 3 % pronto-pago clientes",
                       "Refinanciar deuda en línea revolvente",
                       "Reducir inventario a 15 % del COGS"
                   ]}),
                   "Línea de Acción Gerencial", "")
ws_d.column_dimensions["A"].width = 60

# ═══════════════════════ 5. GUARDAR LIBRO ══════════════════════════════════════
wb.save(book_path)
print(f"Libro actualizado con éxito ➜ {book_path}")
